#!/usr/bin/env python3
"""
============================================================
MyID Migration Dashboard - Interactive Dash Application
============================================================
PURPOSE: Executive-level dashboard for tracking MyID migration
         SOLO → ACME & Sentry 1.0 → Sentry 2.0

AUDIENCE: Directors, Sr. Managers, Tech Leads, Dev Leads, PM/SMs

DATA SOURCE: MP&A_-_Okta-Based_MyID_Migration_(SOLO to ACME)_v3.xlsx

HOW TO RUN:
    1. Install dependencies: pip install -r requirements.txt
    2. cd "/Users/sam.levine.-nd/Desktop/2026 Projects/MyID_Migration"
    3. python scripts/dashboard_app.py
    4. Open browser to: http://localhost:8050

FEATURES:
    - Real-time data refresh (manual button)
    - Interactive charts with drill-down
    - Historical trend tracking
    - Clickable JIRA links
    - PDF export capability
============================================================
"""

import dash
from dash import dcc, html, dash_table, callback_context
from dash.dependencies import Input, Output, State
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
import os
from datetime import datetime, date
from pathlib import Path

# Try to import openpyxl for Excel reading
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("WARNING: openpyxl not installed. Install with: pip install openpyxl")

# ============================================================
# CONFIGURATION
# ============================================================
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent

# New Excel file path
EXCEL_PATH = PROJECT_DIR / "data" / "input" / "MP&A_-_Okta-Based_MyID_Migration_(SOLO to ACME)_v3.xlsx"
HISTORY_PATH = PROJECT_DIR / "data" / "history" / "migration_history.json"
JIRA_BASE_URL = "https://jira.studio.disney.com/browse/"

# Target deadline
DEADLINE_DATE = date(2026, 2, 27)

# ============================================================
# COLOR SCHEME - Corporate Blue + Traffic Light
# ============================================================
COLORS = {
    # Status colors (Traffic Light)
    'completed': '#059669',      # Green (Done)
    'todo': '#f59e0b',           # Amber/Yellow (To Do)
    'in_progress': '#3b82f6',    # Blue (In Progress)
    'blocked': '#dc2626',        # Red
    'obsolete': '#64748b',       # Slate gray
    'at_risk': '#dc2626',        # Red (critical)
    
    # Corporate blues
    'primary': '#1e40af',        # Deep blue
    'primary_light': '#3b82f6',  # Light blue
    'secondary': '#475569',      # Dark slate
    
    # Backgrounds
    'bg_light': '#f8fafc',
    'bg_card': '#ffffff',
    'border': '#e2e8f0',
    'text': '#334155',
    'text_muted': '#64748b',
}

# ============================================================
# DATA LOADING & PROCESSING
# ============================================================
def load_excel_data():
    """Load and process the Excel data from new format"""
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return []
    
    if not EXCEL_SUPPORT:
        print("ERROR: openpyxl not installed")
        return []
    
    apps = []
    
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        
        # Column mapping based on new format (0-indexed)
        # Row 2 has headers, data starts at row 3
        COL_MAP = {
            'platform_app': 0,      # A: Platform/APP
            'system': 1,            # B: System
            'poc': 2,               # C: POC
            'obsolete': 3,          # D: Obsolete
            'sentry1': 4,           # E: Sentry 1.0
            'sentry2': 5,           # F: Sentry 2.0
            'direct_int': 6,        # G: Direct Int.
            'debut_sentry': 7,      # H: Debut-Sentry
            'auth_status': 8,       # I: Auth Sys Status
            'solo': 9,              # J: SOLO
            'acme': 10,             # K: ACME
            'migration_status': 11, # L: MyID Migration (to ACME) Status
            'infra_req': 12,        # M: Infra Access Req'd
            'due_date': 13,         # N: Due Date
            'future_needed': 14,    # O: Needed in Future
            'jira': 15,             # P: JIRA Ticket
            'notes': 16,            # Q: Notes
        }
        
        def get_cell(row, col_key):
            """Get cell value safely"""
            idx = COL_MAP[col_key]
            if idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else ''
            return ''
        
        def is_yes(val):
            """Check if value means 'yes'"""
            return val in ['✓', '✔', 'yes', 'Yes', 'YES', 'y', 'Y', 'true', 'True', '1']
        
        def is_no(val):
            """Check if value means 'no'"""
            return val in ['x', 'X', 'no', 'No', 'NO', 'n', 'N', 'false', 'False', '0']
        
        def is_na(val):
            """Check if value is N/A"""
            return val.lower() in ['n/a', 'na', 'not applicable', ''] if val else True
        
        # Process data rows (starting from row 3, index 2)
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            row_data = list(row)
            
            platform_app = get_cell(row_data, 'platform_app')
            if not platform_app:
                continue  # Skip empty rows
            
            system = get_cell(row_data, 'system')
            poc = get_cell(row_data, 'poc')
            obsolete = get_cell(row_data, 'obsolete')
            sentry1 = get_cell(row_data, 'sentry1')
            sentry2 = get_cell(row_data, 'sentry2')
            direct_int = get_cell(row_data, 'direct_int')
            debut_sentry = get_cell(row_data, 'debut_sentry')
            auth_status = get_cell(row_data, 'auth_status')
            solo = get_cell(row_data, 'solo')
            acme = get_cell(row_data, 'acme')
            migration_status = get_cell(row_data, 'migration_status')
            infra_req = get_cell(row_data, 'infra_req')
            due_date_raw = get_cell(row_data, 'due_date')
            future_needed = get_cell(row_data, 'future_needed')
            jira = get_cell(row_data, 'jira')
            notes = get_cell(row_data, 'notes')
            
            # Parse due date
            due_date = ''
            if due_date_raw and due_date_raw not in ['n/a', 'TBD', 'None']:
                try:
                    if '2025' in due_date_raw or '2026' in due_date_raw:
                        # Already formatted
                        due_date = due_date_raw.split(' ')[0]  # Remove time part
                except:
                    due_date = due_date_raw
            
            # Determine environment from app name
            env = ''
            app_lower = platform_app.lower()
            if '-int' in app_lower or '_int' in app_lower:
                env = 'INT'
            elif '-prd' in app_lower or '-prod' in app_lower or '_prd' in app_lower:
                env = 'PRD'
            elif '-stg' in app_lower or '-stage' in app_lower or '_stg' in app_lower:
                env = 'STG'
            elif '-qat' in app_lower or '_qat' in app_lower:
                env = 'QAT'
            elif 'localhost' in app_lower:
                env = 'LOCAL'
            
            # Determine current auth system
            current_auth = ''
            if is_yes(sentry1) and not is_yes(sentry2):
                current_auth = 'Sentry 1.0'
            elif is_yes(sentry2):
                current_auth = 'Sentry 2.0'
            elif is_yes(debut_sentry):
                current_auth = 'Debut Sentry'
            elif is_yes(direct_int):
                current_auth = 'Direct Integration'
            elif is_yes(solo):
                current_auth = 'SOLO'
            elif is_yes(acme):
                current_auth = 'ACME'
            else:
                current_auth = 'Unknown'
            
            # Normalize migration status based on legend:
            # Done = Complete
            # To Do = Not Started
            # Obsolete = No Migration Required
            # n/a = Not Applicable
            acme_status = migration_status
            if acme_status.lower() == 'done':
                acme_status = 'Done'
            elif acme_status.lower() == 'to do':
                acme_status = 'To Do'
            elif acme_status.lower() == 'in progress':
                acme_status = 'In Progress'
            elif acme_status.lower() == 'obsolete':
                acme_status = 'No Migration Required'
            elif acme_status.lower() in ['n/a', '']:
                acme_status = 'N/A'
            
            # Systems marked "(Not in use)" are treated as obsolete
            is_system_not_in_use = '(not in use)' in system.lower()
            
            # Determine Sentry migration status
            sentry_status = auth_status
            if sentry_status.lower() == 'done':
                sentry_status = 'Done'
            elif sentry_status.lower() == 'in progress':
                sentry_status = 'In Progress'
            elif sentry_status.lower() in ['obsolete', 'n/a', '']:
                sentry_status = 'N/A'
            
            # Determine if at risk:
            # Due Date = 2/27/26 AND Status is NOT Done
            # (includes To Do AND In Progress items)
            at_risk = 'No'
            if due_date and '2026-02-27' in due_date and acme_status not in ['Done', 'No Migration Required', 'N/A']:
                at_risk = 'Yes'
            
            # Check for blocker
            blocker = ''
            if 'block' in notes.lower() or 'waiting' in notes.lower():
                # Extract blocker info from notes
                blocker = notes[:100] if len(notes) > 100 else notes
            
            # Obsolete if: column D = ✓ OR system is "(Not in use)"
            is_obsolete = is_yes(obsolete) or is_system_not_in_use
            
            apps.append({
                'Platform/App': platform_app,
                'Environment': env,
                'System': system.replace(' (Not in use)', '').replace('(Not in use)', '').strip(),
                'System_Full': system,  # Keep original for display
                'POC': poc,
                'Obsolete': is_obsolete,
                'Current Auth System': current_auth,
                'ACME Migration Status': acme_status,
                'Sentry Migration Status': sentry_status,
                'On SOLO': is_yes(solo),
                'On ACME': is_yes(acme),
                'On Sentry 1.0': is_yes(sentry1),
                'On Sentry 2.0': is_yes(sentry2),
                'Due Date': due_date,
                'Future Needed': is_yes(future_needed),
                'At Risk': at_risk,
                'JIRA Ticket': jira,
                'Blocker': blocker,
                'Notes': notes,
                'Infra Required': infra_req,
            })
        
        print(f"Loaded {len(apps)} applications from Excel")
        return apps
        
    except Exception as e:
        print(f"Error loading Excel: {e}")
        import traceback
        traceback.print_exc()
        return []


def calculate_metrics(apps):
    """Calculate KPI metrics from app list"""
    if not apps:
        return {}
    
    total = len(apps)
    
    # Filter out obsolete apps for main metrics
    # Obsolete = column D checked OR system "(Not in use)"
    active_apps = [a for a in apps if not a.get('Obsolete', False)]
    
    # ACME migration metrics
    acme_done = len([a for a in active_apps if a['ACME Migration Status'] == 'Done'])
    acme_todo = len([a for a in active_apps if a['ACME Migration Status'] == 'To Do'])
    acme_in_progress = len([a for a in active_apps if a['ACME Migration Status'] == 'In Progress'])
    acme_no_migration = len([a for a in active_apps if a['ACME Migration Status'] == 'No Migration Required'])
    acme_na = len([a for a in active_apps if a['ACME Migration Status'] == 'N/A'])
    
    # Future needed items also count toward To Do
    future_needed = len([a for a in active_apps if a.get('Future Needed', False) and a['ACME Migration Status'] != 'Done'])
    
    # Total requiring migration = Done + To Do + In Progress (excludes N/A and No Migration Required)
    acme_requiring = acme_done + acme_todo + acme_in_progress
    acme_pct = round((acme_done / acme_requiring * 100), 1) if acme_requiring > 0 else 0
    
    # Sentry/Auth System migration metrics
    sentry_done = len([a for a in active_apps if a['Sentry Migration Status'] == 'Done'])
    sentry_in_progress = len([a for a in active_apps if a['Sentry Migration Status'] == 'In Progress'])
    sentry_na = len([a for a in active_apps if a['Sentry Migration Status'] == 'N/A'])
    
    sentry_requiring = sentry_done + sentry_in_progress
    sentry_pct = round((sentry_done / sentry_requiring * 100), 1) if sentry_requiring > 0 else 0
    
    # At risk: Due 2/27/26 AND NOT Done (includes To Do and In Progress)
    at_risk = len([a for a in active_apps if a['At Risk'] == 'Yes'])
    
    # Blocked (check notes for blocker mentions)
    blocked = len([a for a in active_apps if a.get('Blocker', '')])
    
    # Obsolete/No Migration Required count
    obsolete = len([a for a in apps if a.get('Obsolete', False)])
    
    # Days to deadline
    days_to_deadline = (DEADLINE_DATE - date.today()).days
    
    return {
        'total': total,
        'active': len(active_apps),
        'obsolete': obsolete,
        'acme_done': acme_done,
        'acme_todo': acme_todo,
        'acme_in_progress': acme_in_progress,
        'acme_no_migration': acme_no_migration,
        'acme_na': acme_na,
        'acme_pct': acme_pct,
        'future_needed': future_needed,
        'sentry_done': sentry_done,
        'sentry_in_progress': sentry_in_progress,
        'sentry_na': sentry_na,
        'sentry_pct': sentry_pct,
        'at_risk': at_risk,
        'blocked': blocked,
        'days_to_deadline': days_to_deadline,
    }


def save_history_snapshot(metrics):
    """Save current metrics to history JSON for trend tracking"""
    HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    
    history = []
    if HISTORY_PATH.exists():
        try:
            with open(HISTORY_PATH, 'r') as f:
                history = json.load(f)
        except:
            history = []
    
    snapshot = {
        'date': datetime.now().strftime('%Y-%m-%d'),
        'timestamp': datetime.now().isoformat(),
        **metrics
    }
    
    # Only add if date changed or first entry
    if not history or history[-1]['date'] != snapshot['date']:
        history.append(snapshot)
        with open(HISTORY_PATH, 'w') as f:
            json.dump(history, f, indent=2)
    
    return history


def load_history():
    """Load historical data for trends"""
    if HISTORY_PATH.exists():
        try:
            with open(HISTORY_PATH, 'r') as f:
                return json.load(f)
        except:
            return []
    return []


# ============================================================
# DASH APP INITIALIZATION
# ============================================================
app = dash.Dash(
    __name__,
    title="MyID Migration Dashboard",
    update_title="Updating...",
    suppress_callback_exceptions=True
)

# For gunicorn deployment
server = app.server

# ============================================================
# LAYOUT COMPONENTS
# ============================================================

def create_kpi_card(title, value, subtitle="", color=COLORS['primary'], icon=""):
    """Create a KPI metric card"""
    return html.Div([
        html.Div([
            html.Span(icon, style={'fontSize': '24px', 'marginRight': '10px'}),
            html.Span(title, style={'fontSize': '14px', 'color': COLORS['text_muted'], 'fontWeight': '500'})
        ], style={'marginBottom': '8px'}),
        html.Div(value, style={
            'fontSize': '36px',
            'fontWeight': '700',
            'color': color,
            'lineHeight': '1.2'
        }),
        html.Div(subtitle, style={'fontSize': '12px', 'color': COLORS['text_muted'], 'marginTop': '4px'})
    ], style={
        'background': COLORS['bg_card'],
        'padding': '20px',
        'borderRadius': '12px',
        'boxShadow': '0 2px 8px rgba(0,0,0,0.08)',
        'border': f'1px solid {COLORS["border"]}',
        'flex': '1',
        'minWidth': '180px'
    })


def create_section_header(title, badge_text="", badge_color=COLORS['primary']):
    """Create a section header with optional badge"""
    return html.Div([
        html.H3(title, style={
            'margin': '0',
            'fontSize': '16px',
            'fontWeight': '600',
            'color': COLORS['text']
        }),
        html.Span(badge_text, style={
            'background': badge_color,
            'color': 'white',
            'padding': '4px 12px',
            'borderRadius': '20px',
            'fontSize': '12px',
            'fontWeight': '600'
        }) if badge_text else None
    ], style={
        'display': 'flex',
        'justifyContent': 'space-between',
        'alignItems': 'center',
        'padding': '15px 20px',
        'background': COLORS['bg_light'],
        'borderBottom': f'1px solid {COLORS["border"]}'
    })


# ============================================================
# MAIN LAYOUT
# ============================================================
app.layout = html.Div([
    # Data store
    dcc.Store(id='data-store'),
    dcc.Store(id='metrics-store'),
    
    # Header
    html.Div([
        html.Div([
            html.H1("MyID Migration Dashboard", style={
                'margin': '0',
                'fontSize': '28px',
                'fontWeight': '600'
            }),
            html.P("SOLO → ACME & Sentry 1.0 → Sentry 2.0 Migration Tracking", style={
                'margin': '5px 0 0 0',
                'opacity': '0.9',
                'fontSize': '14px'
            }),
            html.P("Source: MP&A Okta-Based MyID Migration v3", style={
                'margin': '3px 0 0 0',
                'opacity': '0.7',
                'fontSize': '12px'
            }),
        ]),
        html.Div([
            html.Button("🔄 Refresh Data", id='refresh-btn', n_clicks=0, style={
                'background': 'rgba(255,255,255,0.2)',
                'border': '1px solid rgba(255,255,255,0.3)',
                'color': 'white',
                'padding': '10px 20px',
                'borderRadius': '8px',
                'cursor': 'pointer',
                'fontSize': '14px',
                'fontWeight': '500',
                'marginRight': '10px'
            }),
            html.Button("📄 Export PDF", id='export-btn', n_clicks=0, style={
                'background': 'rgba(255,255,255,0.2)',
                'border': '1px solid rgba(255,255,255,0.3)',
                'color': 'white',
                'padding': '10px 20px',
                'borderRadius': '8px',
                'cursor': 'pointer',
                'fontSize': '14px',
                'fontWeight': '500'
            }),
        ], style={'display': 'flex', 'alignItems': 'center'}),
    ], style={
        'background': f'linear-gradient(135deg, {COLORS["primary"]} 0%, {COLORS["secondary"]} 100%)',
        'padding': '25px 30px',
        'color': 'white',
        'display': 'flex',
        'justifyContent': 'space-between',
        'alignItems': 'center',
        'marginBottom': '25px',
        'borderRadius': '12px'
    }),
    
    # Last updated timestamp
    html.Div(id='last-updated', style={
        'textAlign': 'right',
        'fontSize': '12px',
        'color': COLORS['text_muted'],
        'marginBottom': '15px',
        'marginTop': '-15px'
    }),
    
    # ============================================================
    # KPI HERO SECTION
    # ============================================================
    html.Div(id='kpi-section', style={
        'display': 'flex',
        'gap': '20px',
        'marginBottom': '25px',
        'flexWrap': 'wrap'
    }),
    
    # ============================================================
    # AT RISK ALERT (Top Priority)
    # ============================================================
    html.Div(id='at-risk-alert', style={'marginBottom': '25px'}),
    
    # ============================================================
    # FILTER SECTION
    # ============================================================
    html.Div([
        html.Div([
            html.Label("Filter by System:", style={'fontWeight': '500', 'marginBottom': '5px', 'display': 'block'}),
            dcc.Dropdown(
                id='filter-system',
                options=[],
                value=None,
                placeholder="All Systems",
                style={'width': '200px'}
            )
        ], style={'marginRight': '20px'}),
        html.Div([
            html.Label("Filter by Environment:", style={'fontWeight': '500', 'marginBottom': '5px', 'display': 'block'}),
            dcc.Dropdown(
                id='filter-env',
                options=[],
                value=None,
                placeholder="All Environments",
                style={'width': '200px'}
            )
        ], style={'marginRight': '20px'}),
        html.Div([
            html.Label("Filter by POC:", style={'fontWeight': '500', 'marginBottom': '5px', 'display': 'block'}),
            dcc.Dropdown(
                id='filter-poc',
                options=[],
                value=None,
                placeholder="All POCs",
                style={'width': '200px'}
            )
        ], style={'marginRight': '20px'}),
        html.Div([
            html.Label("Filter by Auth System:", style={'fontWeight': '500', 'marginBottom': '5px', 'display': 'block'}),
            dcc.Dropdown(
                id='filter-auth',
                options=[],
                value=None,
                placeholder="All Auth Systems",
                style={'width': '200px'}
            )
        ]),
        html.Div([
            html.Label("Show Obsolete:", style={'fontWeight': '500', 'marginBottom': '5px', 'display': 'block'}),
            dcc.Checklist(
                id='show-obsolete',
                options=[{'label': ' Include obsolete apps', 'value': 'yes'}],
                value=[],
                style={'fontSize': '13px'}
            )
        ], style={'marginLeft': '20px'}),
    ], style={
        'display': 'flex',
        'flexWrap': 'wrap',
        'gap': '15px',
        'padding': '20px',
        'background': COLORS['bg_card'],
        'borderRadius': '12px',
        'border': f'1px solid {COLORS["border"]}',
        'marginBottom': '25px',
        'alignItems': 'flex-end'
    }),
    
    # ============================================================
    # CHARTS ROW 1: Status Overview
    # ============================================================
    html.Div([
        # ACME Status Donut
        html.Div([
            create_section_header("ACME Migration Status (SOLO → ACME)"),
            dcc.Graph(id='acme-donut', config={'displayModeBar': False}, style={'height': '300px'})
        ], style={
            'background': COLORS['bg_card'],
            'borderRadius': '12px',
            'border': f'1px solid {COLORS["border"]}',
            'overflow': 'hidden',
            'flex': '1'
        }),
        
        # Sentry Status Donut
        html.Div([
            create_section_header("Auth System Status"),
            dcc.Graph(id='sentry-donut', config={'displayModeBar': False}, style={'height': '300px'})
        ], style={
            'background': COLORS['bg_card'],
            'borderRadius': '12px',
            'border': f'1px solid {COLORS["border"]}',
            'overflow': 'hidden',
            'flex': '1'
        }),
        
        # Progress Bars
        html.Div([
            create_section_header("Overall Progress"),
            html.Div(id='progress-bars', style={'padding': '20px'})
        ], style={
            'background': COLORS['bg_card'],
            'borderRadius': '12px',
            'border': f'1px solid {COLORS["border"]}',
            'overflow': 'hidden',
            'flex': '1'
        }),
    ], style={'display': 'flex', 'gap': '20px', 'marginBottom': '25px'}),
    
    # ============================================================
    # CHARTS ROW 2: Breakdowns
    # ============================================================
    html.Div([
        # System Breakdown Bar Chart
        html.Div([
            create_section_header("Migration by System"),
            dcc.Graph(id='system-bar', config={'displayModeBar': False}, style={'height': '350px'})
        ], style={
            'background': COLORS['bg_card'],
            'borderRadius': '12px',
            'border': f'1px solid {COLORS["border"]}',
            'overflow': 'hidden',
            'flex': '1.5'
        }),
        
        # POC Workload
        html.Div([
            create_section_header("Work by POC"),
            dcc.Graph(id='poc-bar', config={'displayModeBar': False}, style={'height': '350px'})
        ], style={
            'background': COLORS['bg_card'],
            'borderRadius': '12px',
            'border': f'1px solid {COLORS["border"]}',
            'overflow': 'hidden',
            'flex': '1'
        }),
    ], style={'display': 'flex', 'gap': '20px', 'marginBottom': '25px'}),
    
    # ============================================================
    # HISTORICAL TRENDS
    # ============================================================
    html.Div([
        create_section_header("Migration Progress Over Time", "Historical Trends"),
        dcc.Graph(id='trend-chart', config={'displayModeBar': False}, style={'height': '300px', 'padding': '10px'})
    ], style={
        'background': COLORS['bg_card'],
        'borderRadius': '12px',
        'border': f'1px solid {COLORS["border"]}',
        'overflow': 'hidden',
        'marginBottom': '25px'
    }),
    
    # ============================================================
    # AT RISK TABLE (Prominent - for 2/27/26 deadline)
    # ============================================================
    html.Div([
        create_section_header("⚠️ At Risk for 2/27/26 Deadline", badge_color=COLORS['at_risk']),
        html.Div(id='at-risk-table', style={'padding': '0'})
    ], style={
        'background': COLORS['bg_card'],
        'borderRadius': '12px',
        'border': f'2px solid {COLORS["at_risk"]}',
        'overflow': 'hidden',
        'marginBottom': '25px'
    }),
    
    # ============================================================
    # FULL DATA TABLE
    # ============================================================
    html.Div([
        create_section_header("Complete Migration Tracker"),
        html.Div(id='full-table', style={'padding': '0', 'overflowX': 'auto'})
    ], style={
        'background': COLORS['bg_card'],
        'borderRadius': '12px',
        'border': f'1px solid {COLORS["border"]}',
        'overflow': 'hidden',
        'marginBottom': '25px'
    }),
    
    # ============================================================
    # FUTURE WORK SECTION
    # ============================================================
    html.Div([
        create_section_header("📅 Future Migrations Needed", badge_color=COLORS['todo']),
        html.Div(id='future-table', style={'padding': '0'})
    ], style={
        'background': COLORS['bg_card'],
        'borderRadius': '12px',
        'border': f'1px solid {COLORS["border"]}',
        'overflow': 'hidden',
        'marginBottom': '25px'
    }),
    
    # ============================================================
    # OBSOLETE APPS (Excluded from metrics)
    # ============================================================
    html.Div([
        create_section_header("📦 Obsolete / No Migration Required (Excluded from Metrics)", badge_color=COLORS['obsolete']),
        html.Div(id='obsolete-table', style={'padding': '0'})
    ], style={
        'background': COLORS['bg_card'],
        'borderRadius': '12px',
        'border': f'1px solid {COLORS["border"]}',
        'overflow': 'hidden',
        'marginBottom': '25px',
        'opacity': '0.85'
    }),
    
    # Footer
    html.Div([
        html.P([
            "MyID Migration Dashboard • Data source: ",
            html.Code(EXCEL_PATH.name),
            " • Built with Dash/Plotly"
        ], style={'margin': '0'})
    ], style={
        'textAlign': 'center',
        'padding': '20px',
        'color': COLORS['text_muted'],
        'fontSize': '12px'
    }),
    
    # PDF Export Component (hidden)
    dcc.Download(id='download-pdf'),
    
], style={
    'fontFamily': '-apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif',
    'background': '#f0f4f8',
    'minHeight': '100vh',
    'padding': '20px',
    'color': COLORS['text']
})


# ============================================================
# CALLBACKS
# ============================================================

@app.callback(
    [Output('data-store', 'data'),
     Output('metrics-store', 'data'),
     Output('last-updated', 'children'),
     Output('filter-system', 'options'),
     Output('filter-env', 'options'),
     Output('filter-poc', 'options'),
     Output('filter-auth', 'options')],
    [Input('refresh-btn', 'n_clicks')],
    prevent_initial_call=False
)
def load_and_store_data(n_clicks):
    """Load data and populate filters"""
    apps = load_excel_data()
    metrics = calculate_metrics(apps)
    
    # Save historical snapshot
    if metrics:
        save_history_snapshot(metrics)
    
    # Get unique values for filters (excluding obsolete for cleaner lists)
    active_apps = [a for a in apps if not a.get('Obsolete', False)]
    
    systems = [{'label': s, 'value': s} for s in sorted(set(a['System'] for a in active_apps if a['System']))]
    envs = [{'label': e, 'value': e} for e in sorted(set(a['Environment'] for a in active_apps if a['Environment']))]
    pocs = [{'label': p, 'value': p} for p in sorted(set(a['POC'] for a in active_apps if a['POC']))]
    auths = [{'label': a, 'value': a} for a in sorted(set(app['Current Auth System'] for app in active_apps if app['Current Auth System']))]
    
    timestamp = f"Last updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    return json.dumps(apps), json.dumps(metrics), timestamp, systems, envs, pocs, auths


@app.callback(
    Output('kpi-section', 'children'),
    [Input('metrics-store', 'data')]
)
def update_kpis(metrics_json):
    """Update KPI hero section"""
    if not metrics_json:
        return []
    
    metrics = json.loads(metrics_json)
    
    days = metrics.get('days_to_deadline', 0)
    days_color = COLORS['completed'] if days > 30 else (COLORS['todo'] if days > 14 else COLORS['at_risk'])
    
    acme_done = metrics.get('acme_done', 0)
    acme_total = acme_done + metrics.get('acme_todo', 0) + metrics.get('acme_in_progress', 0)
    
    return [
        create_kpi_card(
            "ACME Migration Progress",
            f"{metrics.get('acme_pct', 0)}%",
            f"{acme_done} of {acme_total} completed",
            COLORS['primary'],
            "📊"
        ),
        create_kpi_card(
            "At Risk Items",
            str(metrics.get('at_risk', 0)),
            "Due 2/27/26 - Not Done",
            COLORS['at_risk'] if metrics.get('at_risk', 0) > 0 else COLORS['completed'],
            "⚠️"
        ),
        create_kpi_card(
            "To Do",
            str(metrics.get('acme_todo', 0)),
            "Awaiting migration",
            COLORS['todo'] if metrics.get('acme_todo', 0) > 0 else COLORS['completed'],
            "📝"
        ),
        create_kpi_card(
            "Days to Deadline",
            str(days),
            f"Target: Feb 27, 2026",
            days_color,
            "📅"
        ),
    ]


@app.callback(
    Output('at-risk-alert', 'children'),
    [Input('metrics-store', 'data')]
)
def update_at_risk_alert(metrics_json):
    """Show prominent at-risk alert if items exist"""
    if not metrics_json:
        return None
    
    metrics = json.loads(metrics_json)
    at_risk = metrics.get('at_risk', 0)
    
    if at_risk == 0:
        return None
    
    return html.Div([
        html.Div([
            html.Span("🚨", style={'fontSize': '24px', 'marginRight': '15px'}),
            html.Div([
                html.Strong(f"CRITICAL: {at_risk} items at risk for 2/27/26 deadline", style={'fontSize': '16px'}),
                html.P("These items have a due date of 2/27/26 and are not yet completed. Immediate attention required.", 
                       style={'margin': '5px 0 0 0', 'opacity': '0.9', 'fontSize': '13px'})
            ])
        ], style={'display': 'flex', 'alignItems': 'center'})
    ], style={
        'background': f'linear-gradient(90deg, {COLORS["at_risk"]}, #b91c1c)',
        'color': 'white',
        'padding': '20px 25px',
        'borderRadius': '12px',
        'boxShadow': '0 4px 12px rgba(220, 38, 38, 0.3)'
    })


def filter_apps(apps, system, env, poc, auth, show_obsolete):
    """Apply filters to app list"""
    filtered = apps
    
    if not show_obsolete or 'yes' not in show_obsolete:
        filtered = [a for a in filtered if not a.get('Obsolete', False) and a['ACME Migration Status'] != 'Obsolete']
    
    if system:
        filtered = [a for a in filtered if a['System'] == system]
    if env:
        filtered = [a for a in filtered if a['Environment'] == env]
    if poc:
        filtered = [a for a in filtered if a['POC'] == poc]
    if auth:
        filtered = [a for a in filtered if a['Current Auth System'] == auth]
    
    return filtered


@app.callback(
    Output('acme-donut', 'figure'),
    [Input('data-store', 'data'),
     Input('filter-system', 'value'),
     Input('filter-env', 'value'),
     Input('filter-poc', 'value'),
     Input('filter-auth', 'value'),
     Input('show-obsolete', 'value')]
)
def update_acme_donut(data_json, system, env, poc, auth, show_obsolete):
    """Update ACME status donut chart"""
    if not data_json:
        return go.Figure()
    
    apps = json.loads(data_json)
    apps = filter_apps(apps, system, env, poc, auth, show_obsolete)
    
    # Count by status
    status_counts = {}
    for app in apps:
        status = app['ACME Migration Status']
        status_counts[status] = status_counts.get(status, 0) + 1
    
    colors_map = {
        'Done': COLORS['completed'],
        'To Do': COLORS['todo'],
        'In Progress': COLORS['in_progress'],
        'No Migration Required': COLORS['obsolete'],
        'Obsolete': COLORS['obsolete'],
        'N/A': COLORS['text_muted']
    }
    
    labels = list(status_counts.keys())
    values = list(status_counts.values())
    colors = [colors_map.get(s, COLORS['text_muted']) for s in labels]
    
    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.6,
        marker_colors=colors,
        textinfo='label+value',
        textposition='outside',
        hovertemplate='%{label}: %{value} apps<extra></extra>'
    )])
    
    fig.update_layout(
        showlegend=True,
        legend=dict(orientation='h', yanchor='bottom', y=-0.2),
        margin=dict(t=20, b=60, l=20, r=20),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    
    return fig


@app.callback(
    Output('sentry-donut', 'figure'),
    [Input('data-store', 'data'),
     Input('filter-system', 'value'),
     Input('filter-env', 'value'),
     Input('filter-poc', 'value'),
     Input('filter-auth', 'value'),
     Input('show-obsolete', 'value')]
)
def update_sentry_donut(data_json, system, env, poc, auth, show_obsolete):
    """Update Auth System status donut chart"""
    if not data_json:
        return go.Figure()
    
    apps = json.loads(data_json)
    apps = filter_apps(apps, system, env, poc, auth, show_obsolete)
    
    # Count by sentry status
    status_counts = {}
    for app in apps:
        status = app['Sentry Migration Status']
        status_counts[status] = status_counts.get(status, 0) + 1
    
    colors_map = {
        'Done': COLORS['completed'],
        'In Progress': COLORS['in_progress'],
        'Obsolete': COLORS['obsolete'],
        'N/A': COLORS['text_muted']
    }
    
    labels = list(status_counts.keys())
    values = list(status_counts.values())
    colors = [colors_map.get(s, COLORS['text_muted']) for s in labels]
    
    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.6,
        marker_colors=colors,
        textinfo='label+value',
        textposition='outside',
        hovertemplate='%{label}: %{value} apps<extra></extra>'
    )])
    
    fig.update_layout(
        showlegend=True,
        legend=dict(orientation='h', yanchor='bottom', y=-0.2),
        margin=dict(t=20, b=60, l=20, r=20),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    
    return fig


@app.callback(
    Output('progress-bars', 'children'),
    [Input('metrics-store', 'data')]
)
def update_progress_bars(metrics_json):
    """Update progress bar section"""
    if not metrics_json:
        return []
    
    metrics = json.loads(metrics_json)
    
    def create_progress_bar(label, pct, color):
        return html.Div([
            html.Div([
                html.Span(label, style={'fontWeight': '500', 'color': COLORS['text']}),
                html.Span(f"{pct}%", style={'fontWeight': '600', 'color': color})
            ], style={'display': 'flex', 'justifyContent': 'space-between', 'marginBottom': '8px'}),
            html.Div([
                html.Div(style={
                    'width': f'{pct}%',
                    'height': '100%',
                    'background': f'linear-gradient(90deg, {color}, {color}dd)',
                    'borderRadius': '4px',
                    'transition': 'width 0.5s ease'
                })
            ], style={
                'height': '12px',
                'background': COLORS['border'],
                'borderRadius': '6px',
                'overflow': 'hidden'
            })
        ], style={'marginBottom': '20px'})
    
    return [
        create_progress_bar("ACME Migration (SOLO → ACME)", metrics.get('acme_pct', 0), COLORS['primary']),
        create_progress_bar("Auth System Migration", metrics.get('sentry_pct', 0), '#7c3aed'),
        html.Hr(style={'border': 'none', 'borderTop': f'1px solid {COLORS["border"]}', 'margin': '15px 0'}),
        html.Div([
            html.Div([
                html.Div(str(metrics.get('at_risk', 0)), style={'fontSize': '24px', 'fontWeight': '700', 'color': COLORS['at_risk']}),
                html.Div("At Risk", style={'fontSize': '11px', 'color': COLORS['text_muted']})
            ], style={'textAlign': 'center', 'flex': '1'}),
            html.Div([
                html.Div(str(metrics.get('acme_todo', 0)), style={'fontSize': '24px', 'fontWeight': '700', 'color': COLORS['todo']}),
                html.Div("To Do", style={'fontSize': '11px', 'color': COLORS['text_muted']})
            ], style={'textAlign': 'center', 'flex': '1'}),
            html.Div([
                html.Div(str(metrics.get('obsolete', 0)), style={'fontSize': '24px', 'fontWeight': '700', 'color': COLORS['obsolete']}),
                html.Div("Obsolete", style={'fontSize': '11px', 'color': COLORS['text_muted']})
            ], style={'textAlign': 'center', 'flex': '1'}),
        ], style={'display': 'flex', 'gap': '10px'})
    ]


@app.callback(
    Output('system-bar', 'figure'),
    [Input('data-store', 'data'),
     Input('filter-env', 'value'),
     Input('filter-poc', 'value'),
     Input('filter-auth', 'value'),
     Input('show-obsolete', 'value')]
)
def update_system_bar(data_json, env, poc, auth, show_obsolete):
    """Update system breakdown bar chart"""
    if not data_json:
        return go.Figure()
    
    apps = json.loads(data_json)
    apps = filter_apps(apps, None, env, poc, auth, show_obsolete)
    
    # Group by system and status
    system_status = {}
    for app in apps:
        system = app['System']
        status = app['ACME Migration Status']
        if system not in system_status:
            system_status[system] = {}
        system_status[system][status] = system_status[system].get(status, 0) + 1
    
    systems = sorted(system_status.keys())
    
    fig = go.Figure()
    
    status_colors = {
        'Done': COLORS['completed'],
        'To Do': COLORS['todo'],
        'In Progress': COLORS['in_progress'],
        'No Migration Required': COLORS['obsolete'],
        'Obsolete': COLORS['obsolete'],
        'N/A': COLORS['text_muted']
    }
    
    for status in ['Done', 'In Progress', 'To Do', 'N/A', 'No Migration Required']:
        values = [system_status.get(s, {}).get(status, 0) for s in systems]
        if sum(values) > 0:
            fig.add_trace(go.Bar(
                name=status,
                x=systems,
                y=values,
                marker_color=status_colors.get(status, COLORS['text_muted'])
            ))
    
    fig.update_layout(
        barmode='stack',
        xaxis_title='System',
        yaxis_title='Number of Apps',
        legend=dict(orientation='h', yanchor='bottom', y=-0.3),
        margin=dict(t=20, b=80, l=50, r=20),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(gridcolor=COLORS['border']),
        yaxis=dict(gridcolor=COLORS['border'])
    )
    
    return fig


@app.callback(
    Output('poc-bar', 'figure'),
    [Input('data-store', 'data'),
     Input('filter-system', 'value'),
     Input('filter-env', 'value'),
     Input('filter-auth', 'value'),
     Input('show-obsolete', 'value')]
)
def update_poc_bar(data_json, system, env, auth, show_obsolete):
    """Update POC workload bar chart"""
    if not data_json:
        return go.Figure()
    
    apps = json.loads(data_json)
    apps = filter_apps(apps, system, env, None, auth, show_obsolete)
    
    # Group by POC and status
    poc_status = {}
    for app in apps:
        poc = app['POC']
        if not poc:
            continue
        status = app['ACME Migration Status']
        if poc not in poc_status:
            poc_status[poc] = {}
        poc_status[poc][status] = poc_status[poc].get(status, 0) + 1
    
    pocs = sorted(poc_status.keys())
    
    fig = go.Figure()
    
    status_colors = {
        'Done': COLORS['completed'],
        'To Do': COLORS['todo'],
        'In Progress': COLORS['in_progress']
    }
    
    for status in ['Done', 'In Progress', 'To Do']:
        values = [poc_status.get(p, {}).get(status, 0) for p in pocs]
        if sum(values) > 0:
            fig.add_trace(go.Bar(
                name=status,
                x=pocs,
                y=values,
                marker_color=status_colors.get(status, COLORS['text_muted'])
            ))
    
    fig.update_layout(
        barmode='stack',
        xaxis_title='POC',
        yaxis_title='Apps',
        legend=dict(orientation='h', yanchor='bottom', y=-0.4),
        margin=dict(t=20, b=100, l=50, r=20),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(gridcolor=COLORS['border'], tickangle=45),
        yaxis=dict(gridcolor=COLORS['border'])
    )
    
    return fig


@app.callback(
    Output('trend-chart', 'figure'),
    [Input('refresh-btn', 'n_clicks')]
)
def update_trend_chart(n_clicks):
    """Update historical trend chart"""
    history = load_history()
    
    if not history or len(history) < 2:
        fig = go.Figure()
        fig.add_annotation(
            text="Historical data will appear after multiple snapshots are recorded.\nRefresh the dashboard on different days to build trend data.",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=14, color=COLORS['text_muted']),
            align='center'
        )
        fig.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            xaxis=dict(visible=False),
            yaxis=dict(visible=False)
        )
        return fig
    
    dates = [h['date'] for h in history]
    done = [h.get('acme_done', 0) for h in history]
    todo = [h.get('acme_todo', 0) for h in history]
    at_risk = [h.get('at_risk', 0) for h in history]
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=dates, y=done,
        mode='lines+markers',
        name='Done',
        line=dict(color=COLORS['completed'], width=3),
        marker=dict(size=8)
    ))
    
    fig.add_trace(go.Scatter(
        x=dates, y=todo,
        mode='lines+markers',
        name='To Do',
        line=dict(color=COLORS['todo'], width=3),
        marker=dict(size=8)
    ))
    
    fig.add_trace(go.Scatter(
        x=dates, y=at_risk,
        mode='lines+markers',
        name='At Risk',
        line=dict(color=COLORS['at_risk'], width=2, dash='dot'),
        marker=dict(size=6)
    ))
    
    fig.update_layout(
        xaxis_title='Date',
        yaxis_title='Number of Apps',
        legend=dict(orientation='h', yanchor='bottom', y=-0.3),
        margin=dict(t=20, b=60, l=50, r=20),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(gridcolor=COLORS['border']),
        yaxis=dict(gridcolor=COLORS['border'])
    )
    
    return fig


def create_jira_link(ticket):
    """Create clickable JIRA link"""
    if not ticket or ticket == '':
        return '-'
    
    tickets = [t.strip() for t in str(ticket).split(',')]
    links = []
    for t in tickets:
        if t:
            links.append(html.A(t, href=f"{JIRA_BASE_URL}{t}", target="_blank", 
                               style={'color': COLORS['primary_light'], 'marginRight': '5px'}))
    
    return html.Span(links) if links else '-'


def status_badge(status):
    """Create status badge HTML"""
    colors = {
        'Done': (COLORS['completed'], '#d1fae5'),
        'To Do': ('#b45309', '#fef3c7'),
        'In Progress': (COLORS['in_progress'], '#dbeafe'),
        'No Migration Required': (COLORS['obsolete'], '#f1f5f9'),
        'Obsolete': (COLORS['obsolete'], COLORS['bg_light']),
        'N/A': (COLORS['text_muted'], COLORS['bg_light'])
    }
    fg, bg = colors.get(status, (COLORS['text_muted'], COLORS['bg_light']))
    return html.Span(status, style={
        'background': bg,
        'color': fg,
        'padding': '3px 10px',
        'borderRadius': '12px',
        'fontSize': '11px',
        'fontWeight': '500'
    })


@app.callback(
    Output('at-risk-table', 'children'),
    [Input('data-store', 'data')]
)
def update_at_risk_table(data_json):
    """Update at-risk items table"""
    if not data_json:
        return html.P("No data loaded", style={'padding': '20px', 'color': COLORS['text_muted']})
    
    apps = json.loads(data_json)
    at_risk_apps = [a for a in apps if a['At Risk'] == 'Yes']
    
    if not at_risk_apps:
        return html.Div([
            html.Span("✓", style={'fontSize': '24px', 'color': COLORS['completed'], 'marginRight': '10px'}),
            html.Span("No items at risk for 2/27/26 deadline", style={'color': COLORS['completed'], 'fontWeight': '500'})
        ], style={'padding': '30px', 'textAlign': 'center'})
    
    rows = []
    for app in at_risk_apps:
        rows.append(html.Tr([
            html.Td(app['Platform/App'], style={'fontWeight': '500'}),
            html.Td(app['Environment']),
            html.Td(app['System']),
            html.Td(app['POC']),
            html.Td(status_badge(app['ACME Migration Status'])),
            html.Td(app['Due Date']),
            html.Td(create_jira_link(app['JIRA Ticket'])),
            html.Td(app['Notes'][:60] + '...' if len(str(app['Notes'])) > 60 else app['Notes'], 
                   title=app['Notes'],
                   style={'color': COLORS['text_muted'], 'fontSize': '12px', 'maxWidth': '300px'}),
        ]))
    
    return html.Table([
        html.Thead(html.Tr([
            html.Th("Platform/App"),
            html.Th("Env"),
            html.Th("System"),
            html.Th("POC"),
            html.Th("Status"),
            html.Th("Due Date"),
            html.Th("JIRA"),
            html.Th("Notes"),
        ], style={'background': COLORS['bg_light']})),
        html.Tbody(rows)
    ], style={'width': '100%', 'borderCollapse': 'collapse', 'fontSize': '13px'})


@app.callback(
    Output('full-table', 'children'),
    [Input('data-store', 'data'),
     Input('filter-system', 'value'),
     Input('filter-env', 'value'),
     Input('filter-poc', 'value'),
     Input('filter-auth', 'value'),
     Input('show-obsolete', 'value')]
)
def update_full_table(data_json, system, env, poc, auth, show_obsolete):
    """Update full data table with filters"""
    if not data_json:
        return html.P("No data loaded", style={'padding': '20px', 'color': COLORS['text_muted']})
    
    apps = json.loads(data_json)
    apps = filter_apps(apps, system, env, poc, auth, show_obsolete)
    
    rows = []
    for app in apps:
        rows.append(html.Tr([
            html.Td(app['Platform/App'], style={'fontWeight': '500'}),
            html.Td(app['Environment']),
            html.Td(app['System']),
            html.Td(app['POC']),
            html.Td(app['Current Auth System'], style={'fontSize': '11px'}),
            html.Td(status_badge(app['ACME Migration Status'])),
            html.Td(status_badge(app['Sentry Migration Status'])),
            html.Td(create_jira_link(app['JIRA Ticket'])),
            html.Td(
                "⚠️" if app['At Risk'] == 'Yes' else ("📅" if app.get('Future Needed') else ""),
                style={'textAlign': 'center'}
            ),
        ]))
    
    return html.Table([
        html.Thead(html.Tr([
            html.Th("Platform/App"),
            html.Th("Env"),
            html.Th("System"),
            html.Th("POC"),
            html.Th("Auth System"),
            html.Th("ACME Status"),
            html.Th("Auth Status"),
            html.Th("JIRA"),
            html.Th("Risk", style={'textAlign': 'center'}),
        ], style={'background': COLORS['bg_light']})),
        html.Tbody(rows)
    ], style={'width': '100%', 'borderCollapse': 'collapse', 'fontSize': '12px'})


@app.callback(
    Output('future-table', 'children'),
    [Input('data-store', 'data')]
)
def update_future_table(data_json):
    """Update future work table - apps with 'Future Needed' flag"""
    if not data_json:
        return html.P("No data loaded", style={'padding': '20px', 'color': COLORS['text_muted']})
    
    apps = json.loads(data_json)
    # Future needed AND not already done AND not obsolete
    future_apps = [a for a in apps if a.get('Future Needed', False) 
                   and a['ACME Migration Status'] != 'Done'
                   and not a.get('Obsolete', False)]
    
    if not future_apps:
        return html.Div([
            html.Span("✓", style={'fontSize': '24px', 'color': COLORS['completed'], 'marginRight': '10px'}),
            html.Span("No pending future migrations", style={'color': COLORS['completed'], 'fontWeight': '500'})
        ], style={'padding': '30px', 'textAlign': 'center'})
    
    rows = []
    for app in future_apps:
        rows.append(html.Tr([
            html.Td(app['Platform/App'], style={'fontWeight': '500'}),
            html.Td(app['Environment']),
            html.Td(app['System']),
            html.Td(app['POC']),
            html.Td(status_badge(app['ACME Migration Status'])),
            html.Td(app['Due Date'] if app['Due Date'] else 'TBD'),
            html.Td(create_jira_link(app['JIRA Ticket'])),
            html.Td(app['Notes'][:60] + '...' if len(str(app['Notes'])) > 60 else app['Notes'], 
                   title=app['Notes'],
                   style={'color': COLORS['text_muted'], 'fontSize': '12px'}),
        ]))
    
    return html.Table([
        html.Thead(html.Tr([
            html.Th("Platform/App"),
            html.Th("Env"),
            html.Th("System"),
            html.Th("POC"),
            html.Th("Status"),
            html.Th("Due Date"),
            html.Th("JIRA"),
            html.Th("Notes"),
        ], style={'background': COLORS['bg_light']})),
        html.Tbody(rows)
    ], style={'width': '100%', 'borderCollapse': 'collapse', 'fontSize': '13px'})


@app.callback(
    Output('obsolete-table', 'children'),
    [Input('data-store', 'data')]
)
def update_obsolete_table(data_json):
    """Update obsolete apps table"""
    if not data_json:
        return html.P("No data loaded", style={'padding': '20px', 'color': COLORS['text_muted']})
    
    apps = json.loads(data_json)
    # Obsolete = column D checked OR system "(Not in use)" OR status "No Migration Required"
    obsolete_apps = [a for a in apps if a.get('Obsolete', False) or a['ACME Migration Status'] == 'No Migration Required']
    
    if not obsolete_apps:
        return html.P("No obsolete apps", style={'padding': '20px', 'color': COLORS['text_muted']})
    
    rows = []
    for app in obsolete_apps:
        rows.append(html.Tr([
            html.Td(app['Platform/App']),
            html.Td(app['System_Full']),
            html.Td(app['POC']),
            html.Td(status_badge(app['ACME Migration Status'])),
            html.Td(app['Notes'][:80] + '...' if len(str(app['Notes'])) > 80 else app['Notes'], 
                   title=app['Notes'],
                   style={'color': COLORS['text_muted'], 'fontSize': '12px'}),
        ], style={'opacity': '0.8'}))
    
    return html.Table([
        html.Thead(html.Tr([
            html.Th("Platform/App"),
            html.Th("System"),
            html.Th("POC"),
            html.Th("Status"),
            html.Th("Notes"),
        ], style={'background': COLORS['bg_light']})),
        html.Tbody(rows)
    ], style={'width': '100%', 'borderCollapse': 'collapse', 'fontSize': '12px'})


# Add CSS for table styling
app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            table { border-collapse: collapse; }
            th, td { 
                padding: 12px 16px; 
                text-align: left; 
                border-bottom: 1px solid #e2e8f0;
            }
            th {
                font-weight: 600;
                text-transform: uppercase;
                font-size: 11px;
                letter-spacing: 0.5px;
                color: #64748b;
            }
            tr:hover { background: #f8fafc; }
            a { text-decoration: none; }
            a:hover { text-decoration: underline; }
            
            /* Print styles for PDF export */
            @media print {
                body { background: white !important; }
                .dash-graph { break-inside: avoid; }
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''


# ============================================================
# MAIN ENTRY POINT
# ============================================================
if __name__ == '__main__':
    print("=" * 60)
    print("MyID Migration Dashboard")
    print("=" * 60)
    print(f"Data source: {EXCEL_PATH}")
    print(f"History file: {HISTORY_PATH}")
    print("")
    print("Starting server...")
    print("Open your browser to: http://localhost:8050")
    print("Press Ctrl+C to stop")
    print("=" * 60)
    
    app.run(debug=True, host='0.0.0.0', port=8050)
